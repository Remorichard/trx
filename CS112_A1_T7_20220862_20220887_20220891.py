#Program: Population application with openpyxl
#Explanation of input : the input is of type xlsx where the user enter the name of xlsx file then program print the data on the user terminal
#Author:Richard Remo Emmanuel, ID:20220862. Did the python program, and test the program 
      #:Ajal Gai Debg Gai, ID: 20220871. Did or write the algorithms for writing the program
      #:Sunday Pul Nyok, ID: 20220891. Did the excel(xlsx file) for runing the program
#Date:2, march,2024
#verion: 0.7
import openpyxl
from openpyxl import workbook, load_workbook

def load_country_data(country):
    try:
        wb = openpyxl.load_workbook(f'{country}.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only = True):
            print(row)
    except FileNotFoundError:
        print(f'Statistic for {country} are not available.')


def main():
    while True:
        print("Country Statistics Application")
        print("1. Enter a country to load its staristics")
        print("2. Exit")

        choice = input ("Enter your choice: ")

        if choice == '1':
            country = input("Enter the country name: ")
            load_country_data(country) 

        elif choice == '2':
            print("Exiting the  application. Goodbye")
            break
        else:
            print("Invalid choice. Please try again")
if __name__ == "__main__":
    main()           
